import streamlit as st
import datetime
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from streamlit_drawable_canvas import st_canvas
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Rejilla OPT Completa", layout="wide")

st.title("📋 Sistema de Registro: Rejilla de Observación de Puesto (OPT)")
st.write("Complete el formulario. Al enviar, se guardará en Drive y se actualizará la base de datos central.")

# ==========================================
# FUNCIONES AUXILIARES (INTERFAZ)
# ==========================================
def hacer_pregunta_estandar(id_pregunta, texto_pregunta, tipo="radio", opciones=["Sí", "No"]):
    col_preg, col_obs = st.columns([2, 1])
    with col_preg:
        if tipo == "radio":
            respuesta = st.radio(texto_pregunta, opciones, key=f"q_{id_pregunta}")
        elif tipo == "texto":
            respuesta = st.text_input(texto_pregunta, key=f"q_{id_pregunta}")
        elif tipo == "area":
            respuesta = st.text_area(texto_pregunta, key=f"q_{id_pregunta}")
        elif tipo == "select":
            respuesta = st.selectbox(texto_pregunta, opciones, key=f"q_{id_pregunta}")
    with col_obs:
        observacion = st.text_area("Observaciones", key=f"obs_{id_pregunta}", height=68, label_visibility="collapsed", placeholder="Comentarios y desvíos...")
    st.write("---")
    return respuesta, observacion

def fila_grilla_interactiva(id_item, texto_item):
    g_tit, g1, g2, g3, g4, g5, g_obs = st.columns([3, 1, 1, 1, 1, 1, 3])
    with g_tit: st.markdown(f"<div style='padding-top: 8px;'>{texto_item}</div>", unsafe_allow_html=True)
    with g1: p1 = st.text_input("1", key=f"g_{id_item}_1", label_visibility="collapsed")
    with g2: p2 = st.text_input("2", key=f"g_{id_item}_2", label_visibility="collapsed")
    with g3: p3 = st.text_input("3", key=f"g_{id_item}_3", label_visibility="collapsed")
    with g4: p4 = st.text_input("4", key=f"g_{id_item}_4", label_visibility="collapsed")
    with g5: p5 = st.text_input("5", key=f"g_{id_item}_5", label_visibility="collapsed")
    with g_obs: obs = st.text_input("Obs", key=f"g_{id_item}_obs", label_visibility="collapsed", placeholder="Notas...")
    st.write("---")
    return [p1, p2, p3, p4, p5], obs

# ==========================================
# CREADOR DEL EXCEL
# ==========================================
def generar_excel_profesional(datos_form):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado OPT"
    ws.views.sheetView[0].showGridLines = False 
    
    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    QUESTION_FILL = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    GRILLA_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                         top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))
    
    ws["A1"] = "REJILLA DE OBSERVACIÓN DE PUESTO (OPT) - RESULTADOS"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
    
    ws["A3"] = "Fecha:"; ws["B3"] = datos_form['Encabezado']['Fecha']
    ws["A4"] = "Observador:"; ws["B4"] = datos_form['Encabezado']['Observador']
    ws["D3"] = "UTE / Equipo:"; ws["E3"] = datos_form['Encabezado']['UTE']
    ws["D4"] = "Puesto / Operario:"; ws["E4"] = datos_form['Encabezado']['Operario']
    for row in range(3, 5):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4).font = Font(bold=True)
    
    ws.cell(row=6, column=1, value="ID").fill = HEADER_FILL
    ws.cell(row=6, column=2, value="Criterio / Pregunta de la Observación").fill = HEADER_FILL
    ws.cell(row=6, column=3, value="Respuesta / Valores Registrados").fill = HEADER_FILL
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=7)
    ws.cell(row=6, column=8, value="Observaciones y Comentarios").fill = HEADER_FILL
    
    for c in range(1, 9):
        cell = ws.cell(row=6, column=c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    r = 7
    for seccion, preguntas in datos_form['Contenido'].items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        s_cell = ws.cell(row=r, column=1, value=seccion)
        s_cell.font = Font(name="Arial", size=11, bold=True, color="1F497D")
        s_cell.fill = SECTION_FILL
        s_cell.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 9): ws.cell(row=r, column=col).border = thin_border
        r += 1
        
        if "3. Diversidad" in seccion:
            ws.cell(row=r, column=2, value="Desglose por ciclos:").font = Font(bold=True, italic=True)
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=2).fill = GRILLA_FILL
            ws.cell(row=r, column=1).fill = GRILLA_FILL
            ws.cell(row=r, column=8).fill = GRILLA_FILL
            for p_idx in range(5):
                pz_cell = ws.cell(row=r, column=3 + p_idx, value=f"Pz {p_idx+1}")
                pz_cell.fill = GRILLA_FILL
                pz_cell.font = Font(bold=True)
                pz_cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 9): ws.cell(row=r, column=col).border = thin_border
            r += 1
        
        for q_id, q_data in preguntas.items():
            ws.cell(row=r, column=1, value=q_id).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
            c_preg = ws.cell(row=r, column=2, value=q_data['texto'])
            c_preg.fill = QUESTION_FILL
            c_preg.alignment = Alignment(wrap_text=True, vertical="center")
            
            if q_data['es_grilla']:
                for p_idx in range(5):
                    ws.cell(row=r, column=3 + p_idx, value=q_data['valores'][p_idx]).alignment = Alignment(horizontal="center", vertical="center")
            else:
                val_cell = ws.cell(row=r, column=3, value=q_data['valores'])
                val_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
                
            c_obs = ws.cell(row=r, column=8, value=q_data['observacion'])
            c_obs.alignment = Alignment(wrap_text=True, vertical="center")
            for col in range(1, 9): ws.cell(row=r, column=col).border = thin_border
            r += 1
            
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    for c in ["C", "D", "E", "F", "G"]: ws.column_dimensions[c].width = 11
    ws.column_dimensions["H"].width = 45
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def subir_excel_a_drive(excel_bytes, nombre_archivo, folder_id):
    credenciales_dict = json.loads(st.secrets["google_credentials"])
    SCOPES = ['https://www.googleapis.com/auth/drive']
    creds = service_account.Credentials.from_service_account_info(credenciales_dict, scopes=SCOPES)
    servicio_drive = build('drive', 'v3', credentials=creds)
    archivo_stream = io.BytesIO(excel_bytes)
    media = MediaIoBaseUpload(archivo_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
    metadatos_archivo = {'name': nombre_archivo, 'parents': [folder_id]}
    archivo_subido = servicio_drive.files().create(body=metadatos_archivo, media_body=media, fields='id, webViewLink', supportsAllDrives=True).execute()
    return archivo_subido.get('webViewLink')

# ==========================================
# INTERFAZ WEB DEL FORMULARIO
# ==========================================
with st.form("formulario_opt_completo"):
    st.header("Información del Relevamiento")
    c1, c2, c3, c4 = st.columns(4)
    with c1: fecha = st.date_input("Fecha", datetime.date.today())
    with c2: observador = st.text_input("Observador*")
    with c3: ute_equipo = st.text_input("UTE / Equipo")
    with c4: puesto_operario = st.text_input("Puesto / Operario*")
    st.divider()

    st.header("1. Preparación de la Observación")
    
    # --- CAMBIOS APLICADOS EN EL MÓDULO 1 ---
    r1_1, o1_1 = hacer_pregunta_estandar("1_1", "1.1. ¿Los estándares están al día y completos (FOS, Estado de Referencia 5S, TEO)?")
    r1_2, o1_2 = hacer_pregunta_estandar("1_2", "1.2. ¿El operario ha seguido las formaciones del TEO? (Indique nivel Skill en observaciones)")
    r1_3, o1_3 = hacer_pregunta_estandar("1_3", "1.3. ¿La FSSE esta al día (vs la última modificación del puesto)?")
    r1_4, o1_4 = hacer_pregunta_estandar("1_4", "1.4. ¿Identificó algún problema de Ergonomía o Seguridad? (Si sí, ¿Cuál? en obs.)")
    r1_5, o1_5 = hacer_pregunta_estandar("1_5", "1.5. ¿Identificó algún problema de calidad reciente? (Si sí, ¿Cuál? en obs.)")
    
    # Estas dos mantienen su formato original
    r1_6, o1_6 = hacer_pregunta_estandar("1_6", "1.6. ¿Cuál es el indicador prioritario a mejorar en la UTE?", tipo="texto")
    r1_7, o1_7 = hacer_pregunta_estandar("1_7", "1.7. Filtro escogido", tipo="select", opciones=["Seguridad", "Calidad", "Costos", "Plazo", "Ergonomía"])

    st.header("2. Observación de respeto de los estándares - Observación de lejos")
    r2_1, o2_1 = hacer_pregunta_estandar("2_1", "2.1. ¿El operario tiene los EPP mencionados en las FOS A/P - Ficha de Seguridad?")
    r2_2, o2_2 = hacer_pregunta_estandar("2_2", "2.2. ¿El puesto está conforme con el estado de referencia 5S y comprende la señalización CSR?")
    r2_3, o2_3 = hacer_pregunta_estandar("2_3", "2.3. ¿El operario respeta la FOS (Orden de las etapas principales)?")
    r2_4, o2_4 = hacer_pregunta_estandar("2_4", "2.4. ¿Las actividades no cíclicas (Cambios, mantenimiento) son realizadas conforme al estándar?")
    r2_5, o2_5 = hacer_pregunta_estandar("2_5", "2.5. ¿Las actividades de calidad frecuenciales (Poka Yoke) son realizadas conforme al estándar?")

    st.header("3. Diversidad")
    v3_1, obs3_1 = fila_grilla_interactiva("3_1", "3.1. Tiempo Operatorio estándar (FOS)")
    v3_2, obs3_2 = fila_grilla_interactiva("3_2", "3.2. Tiempo operatorio medido")
    v3_3, obs3_3 = fila_grilla_interactiva("3_3", "3.3. Tiempo de actividades ok o no ok (+/-5%)")
    st.write("**3.4. No Valor Agregado**")
    v3_4a, obs3_4a = fila_grilla_interactiva("3_4a", "• Número de pasos")
    v3_4b, obs3_4b = fila_grilla_interactiva("3_4b", "• Tomar o depositar intermedio")
    v3_4c, obs3_4c = fila_grilla_interactiva("3_4c", "• Esperas")

    st.header("4. Observación del respeto del estándar - Observación de cerca")
    r4_1, o4_1 = hacer_pregunta_estandar("4_1", "4.1. Si una FOS A ha sido definida, ¿el operario la respeta (ligada a problema/defecto)?")
    r4_2, o4_2 = hacer_pregunta_estandar("4_2", "4.2. ¿Los puntos clave son respetados y apropiados a los problemas de calidad/seguridad?")
    r4_3, o4_3 = hacer_pregunta_estandar("4_3", "4.3. ¿El producto esta conforme a lo requerido (Materias Primas & Producto terminado)?")
    r4_4, o4_4 = hacer_pregunta_estandar("4_4", "4.4. ¿Los embalajes, los útiles, las asistencias y las ayudas son las del estado de referencia?")
    r4_5, o4_5 = hacer_pregunta_estandar("4_5", "4.5. ¿Las piezas son identificadas y los registros de trazabilidad son efectuados correctamente?")
    r4_6, o4_6 = hacer_pregunta_estandar("4_6", "4.6. ¿Los procedimientos de gestión de desechos y reglas de seguridad son respetados?")

    st.header("5. Observación para la mejora del estándar")
    r5_1, o5_1 = hacer_pregunta_estandar("5_1", "Identificar las mejoras y acciones a mediano plazo (Plan UTE/LUP)", tipo="area")

    st.header("6. Síntesis de la observación")
    r6_1, o6_1 = hacer_pregunta_estandar("6_1", "6.1. Intercambio con el operario sobre el respeto del estándar", tipo="area")
    r6_2, o6_2 = hacer_pregunta_estandar("6_2", "6.2. ¿El operario es capaz de nombrar etapas principales, prohibiciones y puntos CSR?", opciones=["Sí", "No", "Parcialmente"])
    r6_3, o6_3 = hacer_pregunta_estandar("6_3", "6.3. ¿Hay algún elemento que debe ser adjuntado al Cuadro de Control?")
    r6_4, o6_4 = hacer_pregunta_estandar("6_4", "6.4. Compartir las mejoras (del operario / del observador)", tipo="area")
    r6_5, o6_5 = hacer_pregunta_estandar("6_5", "6.5. ¿Las mejoras pueden ser transversalizadas?")

    st.header("Acciones Inmediatas Realizadas")
    r_acc1, o_acc1 = hacer_pregunta_estandar("acc_1", "Criterios de la OPT / Desviación identificadas", tipo="area")
    r_acc2, o_acc2 = hacer_pregunta_estandar("acc_2", "Acción inmediata realizada", tipo="area")

    st.header("🎨 Croquis Opcional del Puesto")
    canvas_result = st_canvas(fill_color="rgba(255, 165, 0, 0.3)", stroke_width=2, stroke_color="#FF0000", background_color="#f0f2f6", height=200, width=700, drawing_mode="freedraw", key="canvas_opt")

    btn_enviar = st.form_submit_button("🚀 Procesar Respuestas", type="primary")

# ==========================================
# PROCESAMIENTO AL ENVIAR
# ==========================================
if btn_enviar:
    if not observador or not puesto_operario:
        st.error("❌ Por favor complete los campos obligatorios ('Observador' y 'Puesto / Operario').")
    else:
        # 1. Diccionario completo para el Excel
        datos_totales = {
            'Encabezado': {'Fecha': str(fecha), 'Observador': observador, 'UTE': ute_equipo, 'Operario': puesto_operario},
            'Contenido': {
                '1. Preparación de la Observación': {
                    '1.1': {'texto': 'Los estándares están al día y completos', 'es_grilla': False, 'valores': r1_1, 'observacion': o1_1},
                    '1.2': {'texto': 'Formaciones TEO', 'es_grilla': False, 'valores': r1_2, 'observacion': o1_2},
                    '1.3': {'texto': 'FSSE al día', 'es_grilla': False, 'valores': r1_3, 'observacion': o1_3},
                    '1.4': {'texto': 'Problemas Ergonomía o Seguridad', 'es_grilla': False, 'valores': r1_4, 'observacion': o1_4},
                    '1.5': {'texto': 'Problemas calidad reciente', 'es_grilla': False, 'valores': r1_5, 'observacion': o1_5},
                    '1.6': {'texto': 'Indicador prioritario a mejorar', 'es_grilla': False, 'valores': r1_6, 'observacion': o1_6},
                    '1.7': {'texto': 'Filtro escogido', 'es_grilla': False, 'valores': r1_7, 'observacion': o1_7},
                },
                '2. Observación de respeto de los estándares - Observación de lejos': {
                    '2.1': {'texto': 'Uso de EPP', 'es_grilla': False, 'valores': r2_1, 'observacion': o2_1},
                    '2.2': {'texto': 'Estado 5S / CSR', 'es_grilla': False, 'valores': r2_2, 'observacion': o2_2},
                    '2.3': {'texto': 'Respeta FOS (Orden)', 'es_grilla': False, 'valores': r2_3, 'observacion': o2_3},
                    '2.4': {'texto': 'Actividades no cíclicas', 'es_grilla': False, 'valores': r2_4, 'observacion': o2_4},
                    '2.5': {'texto': 'Actividades calidad frecuenciales', 'es_grilla': False, 'valores': r2_5, 'observacion': o2_5},
                },
                '3. Diversidad': {
                    '3.1': {'texto': 'Tiempo Operatorio estándar (FOS)', 'es_grilla': True, 'valores': v3_1, 'observacion': obs3_1},
                    '3.2': {'texto': 'Tiempo operatorio medido', 'es_grilla': True, 'valores': v3_2, 'observacion': obs3_2},
                    '3.3': {'texto': 'Tiempo de actividades ok o no ok', 'es_grilla': True, 'valores': v3_3, 'observacion': obs3_3},
                    '3.4a': {'texto': '• Número de pasos', 'es_grilla': True, 'valores': v3_4a, 'observacion': obs3_4a},
                    '3.4b': {'texto': '• Tomar/Depositar', 'es_grilla': True, 'valores': v3_4b, 'observacion': obs3_4b},
                    '3.4c': {'texto': '• Esperas', 'es_grilla': True, 'valores': v3_4c, 'observacion': obs3_4c},
                },
                '4. Observación del respeto del estándar - Observación de cerca': {
                    '4.1': {'texto': 'FOS A ligada a problema/defecto', 'es_grilla': False, 'valores': r4_1, 'observacion': o4_1},
                    '4.2': {'texto': 'Puntos clave respetados', 'es_grilla': False, 'valores': r4_2, 'observacion': o4_2},
                    '4.3': {'texto': 'Producto conforme', 'es_grilla': False, 'valores': r4_3, 'observacion': o4_3},
                    '4.4': {'texto': 'Embalajes y útiles de referencia', 'es_grilla': False, 'valores': r4_4, 'observacion': o4_4},
                    '4.5': {'texto': 'Trazabilidad y marcaje', 'es_grilla': False, 'valores': r4_5, 'observacion': o4_5},
                    '4.6': {'texto': 'Gestión de desechos y seguridad', 'es_grilla': False, 'valores': r4_6, 'observacion': o4_6},
                },
                '5. Observación para la mejora del estándar': {
                    '5.1': {'texto': 'Mejoras y acciones a mediano plazo', 'es_grilla': False, 'valores': r5_1, 'observacion': o5_1},
                },
                '6. Síntesis de la observación': {
                    '6.1': {'texto': 'Intercambio con el operario', 'es_grilla': False, 'valores': r6_1, 'observacion': o6_1},
                    '6.2': {'texto': 'Conocimiento etapas y CSR', 'es_grilla': False, 'valores': r6_2, 'observacion': o6_2},
                    '6.3': {'texto': 'Adjuntar al Cuadro de Control', 'es_grilla': False, 'valores': r6_3, 'observacion': o6_3},
                    '6.4': {'texto': 'Compartir mejoras', 'es_grilla': False, 'valores': r6_4, 'observacion': o6_4},
                    '6.5': {'texto': 'Mejoras transversalizadas', 'es_grilla': False, 'valores': r6_5, 'observacion': o6_5},
                },
                'Acciones Inmediatas Realizadas': {
                    'Dev': {'texto': 'Desviación identificada', 'es_grilla': False, 'valores': r_acc1, 'observacion': o_acc1},
                    'Acc': {'texto': 'Acción inmediata realizada', 'es_grilla': False, 'valores': r_acc2, 'observacion': o_acc2},
                }
            }
        }
        
        # 2. Generar Excel y subir a Drive
        excel_bytes = generar_excel_profesional(datos_totales)
        nombre_excel = f"OPT_{puesto_operario.replace(' ', '_')}_{fecha}.xlsx"
        ID_CARPETA_DRIVE = "1Io4SlOlxISQIA0Jrriiz5LF5fWEYM5Ys" 
        link_del_excel = "No se pudo subir"
        
        try:
            link_del_excel = subir_excel_a_drive(excel_bytes, nombre_excel, ID_CARPETA_DRIVE)
            st.success("📁 ¡El Excel se guardó en Google Drive!")
            st.markdown(f"[🔗 Ver Excel en la nube]({link_del_excel})")
        except Exception as e:
            st.error(f"Error en Drive: {e}")
            st.download_button("📥 Descargar Excel Manualmente", data=excel_bytes, file_name=nombre_excel)

        # 3. Empaquetar fila completa para Google Sheets
        def formato_pz(lista_pz):
            return " | ".join([p if p else "-" for p in lista_pz])

        fila_completa = [
            str(fecha), observador, ute_equipo, puesto_operario,
            r1_1, o1_1, r1_2, o1_2, r1_3, o1_3, r1_4, o1_4, r1_5, o1_5, r1_6, o1_6, r1_7, o1_7,
            r2_1, o2_1, r2_2, o2_2, r2_3, o2_3, r2_4, o2_4, r2_5, o2_5,
            formato_pz(v3_1), obs3_1, formato_pz(v3_2), obs3_2, formato_pz(v3_3), obs3_3,
            formato_pz(v3_4a), obs3_4a, formato_pz(v3_4b), obs3_4b, formato_pz(v3_4c), obs3_4c,
            r4_1, o4_1, r4_2, o4_2, r4_3, o4_3, r4_4, o4_4, r4_5, o4_5, r4_6, o4_6,
            r5_1, o5_1,
            r6_1, o6_1, r6_2, o6_2, r6_3, o6_3, r6_4, o6_4, r6_5, o6_5,
            r_acc1, o_acc1, r_acc2, o_acc2,
            link_del_excel 
        ]
        
        # 4. Enviar a Google Sheets
        try:
            credenciales_dict = json.loads(st.secrets["google_credentials"])
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
            client = gspread.authorize(creds)
            
            sheet = client.open_by_key("1S-N8Ok_Q0NmYqQSzKUBpXgW3OvcjnNkrNBsrL5Xd4dY").sheet1 
            sheet.append_row(fila_completa)
            st.success("☁️ ¡Base de datos de Sheets actualizada con éxito!")
        except Exception as e:
            st.error(f"Error en Sheets: {e}")

# ==========================================
# BOTÓN PARA REINICIAR (NUEVA OPT)
# ==========================================
st.write("---")
col_espacio1, col_boton, col_espacio2 = st.columns([1, 2, 1])

with col_boton:
    if st.button("🔄 Crear nueva OPT (Limpiar formulario)", use_container_width=True):
        for key in st.session_state.keys():
            del st.session_state[key]
        st.rerun()
